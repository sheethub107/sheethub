# utils/excel_cleaner.py
import pandas as pd
import io
import re
import datetime as pydt
from .header_detection import detect_header_row

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

# -------------------------
# Core cleaning
# -------------------------
def clean_single_sheet_from_raw(df_raw: pd.DataFrame, header_idx: int) -> pd.DataFrame:
    header = df_raw.iloc[header_idx].fillna("").astype(str).tolist()
    data = df_raw.iloc[header_idx + 1 :].copy()

    if data.empty:
        return pd.DataFrame()

    data.columns = header
    data = (
        data.dropna(how="all")
            .dropna(axis=1, how="all")
            .reset_index(drop=True)
    )

    fixed_cols = []
    for i, c in enumerate(data.columns):
        c = str(c).strip()
        fixed_cols.append(c if c and not c.lower().startswith("unnamed") else f"column_{i}")
    data.columns = fixed_cols

    return data

# -------------------------
# Column standardization
# -------------------------
def standardize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    def snake(s):
        s = str(s).lower().strip()
        s = s.replace("%", " pct").replace("&", " and ")
        s = re.sub(r"[^\w\s]", "", s)
        s = re.sub(r"\s+", "_", s)
        return s.strip("_") or "column"

    seen = {}
    cols = []
    for c in df.columns:
        base = snake(c)
        seen[base] = seen.get(base, -1) + 1
        cols.append(base if seen[base] == 0 else f"{base}_{seen[base]}")
    df.columns = cols
    return df

# -------------------------
# Summary row removal
# -------------------------
def remove_summary_rows(df: pd.DataFrame, keywords):
    if not keywords:
        return df

    keys = [k.lower() for k in keywords]

    def is_summary(row):
        for v in row:
            if pd.isna(v):
                continue
            if any(k in str(v).lower() for k in keys):
                return True
        return False

    mask = df.apply(lambda r: is_summary(r.tolist()), axis=1)
    return df.loc[~mask].reset_index(drop=True)

# -------------------------
# Smart deduplication
# -------------------------
def smart_deduplicate(df: pd.DataFrame, subset):
    if not subset:
        return df.drop_duplicates().reset_index(drop=True)

    subset = [c for c in subset if c in df.columns]
    if not subset:
        return df.drop_duplicates().reset_index(drop=True)

    df["_score"] = df.notna().sum(axis=1)

    df = (
        df.sort_values("_score", ascending=False)
          .drop_duplicates(subset=subset, keep="first")
          .drop(columns="_score")
          .reset_index(drop=True)
    )
    return df

# -------------------------
# Excel helpers
# -------------------------
def _style_header(ws):
    fill = PatternFill("solid", fgColor="FFF2CC")
    bold = Font(bold=True)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in ws[1]:
        c.fill = fill
        c.font = bold
        c.alignment = align

def _freeze(ws):
    ws.freeze_panes = "A2"

def _autofit(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = max(
            10, max(len(str(c.value)) if c.value else 0 for c in col) + 2
        )

def _delete_empty_rows(ws):
    for row in range(ws.max_row, 1, -1):
        is_empty = True
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=row, column=col).value not in (None, ""):
                is_empty = False
                break
        if is_empty:
            ws.delete_rows(row)

# -------------------------
# Main pipeline
# -------------------------
def smart_clean_sheets_from_bytes(
    file_bytes: bytes,
    apply_standardize=False,
    remove_summary=False,
    summary_keywords=None,
    remove_dupes=False,
    dup_subset_map=None,
    drop_missing=False,
):
    cleaned = {}
    all_raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, header=None)

    for sheet, df_raw in all_raw.items():
        if df_raw is None or df_raw.empty:
            cleaned[sheet] = pd.DataFrame()
            continue

        header_idx = detect_header_row(df_raw)
        df = clean_single_sheet_from_raw(df_raw, header_idx)

        if apply_standardize:
            df = standardize_column_names(df)

        if remove_summary:
            df = remove_summary_rows(df, summary_keywords)

        if remove_dupes:
            subset = dup_subset_map.get(sheet) if dup_subset_map else ["employeeid"]
            df = smart_deduplicate(df, subset)

        if drop_missing:
            df = df.dropna().reset_index(drop=True)

        cleaned[sheet] = df

    return cleaned

# -------------------------
# Excel output (FINAL)
# -------------------------
def make_excel_bytes_from_sheets(sheets: dict) -> io.BytesIO:
    out = io.BytesIO()

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_name = sheet_name[:31] if sheet_name else "Sheet1"
            df_copy = df.copy()

            # 🔥 convert datetime → date
            for col in df_copy.columns:
                if pd.api.types.is_datetime64_any_dtype(df_copy[col]):
                    df_copy[col] = df_copy[col].dt.date

            df_copy.to_excel(writer, index=False, sheet_name=safe_name)

            ws = writer.book[safe_name]

            _style_header(ws)
            _freeze(ws)
            _autofit(ws)
            _delete_empty_rows(ws)

    out.seek(0)
    return out
